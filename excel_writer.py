"""
Excel writer: Fills the tracker template with extracted data.
"""

import io
import os
from openpyxl import load_workbook
from config import SHEET_NAMES, DATA_START_ROW, TEMPLATE_FILE


def write_tracker(
    rows: list[dict],
    case_number: int,
    template_path: str | None = None,
) -> bytes:
    """
    Write extracted data rows into the tracker template.

    Args:
        rows: List of row dicts, each mapping column_number -> value.
        case_number: 1, 2, or 3.
        template_path: Path to Tracker_Format.xlsx. Defaults to config value.

    Returns:
        Excel workbook as bytes (for Streamlit download).
    """
    if template_path is None:
        template_path = os.path.join(os.path.dirname(__file__), TEMPLATE_FILE)

    wb = load_workbook(template_path)
    sheet_name = SHEET_NAMES[case_number]
    ws = wb[sheet_name]

    for row_idx, row_data in enumerate(rows):
        excel_row = DATA_START_ROW + row_idx
        for col_num, value in row_data.items():
            # Safety: convert any remaining lists/dicts to strings
            if isinstance(value, (list, tuple)):
                value = ", ".join(str(v) for v in value if v is not None)
            elif isinstance(value, dict):
                value = str(value)
            ws.cell(row=excel_row, column=col_num, value=value)

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
